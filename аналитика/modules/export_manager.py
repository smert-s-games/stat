"""
Модуль для экспорта данных
"""
import csv
import json
import os
from datetime import datetime
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class ExportManager:
    def __init__(self):
        pass
    
    def export_stats_to_csv(self, results, filename=None):
        """Экспорт статистики в CSV"""
        if filename is None:
            filename = f"stats_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                
                # Заголовки
                writer.writerow(['Канал', 'URL', 'Подписчики', 'Просмотры', 'Видео', 'Дата регистрации', 'Страна', 'Статус'])
                
                # Данные
                for result in results:
                    if 'error' not in result:
                        writer.writerow([
                            result.get('channel_name', ''),
                            result.get('url', ''),
                            result.get('subscribers', '0'),
                            result.get('total_views', '0'),
                            result.get('videos_count', '0'),
                            result.get('join_date', ''),
                            result.get('country', ''),
                            'Успешно'
                        ])
                    else:
                        writer.writerow([
                            '',
                            result.get('url', ''),
                            '',
                            '',
                            '',
                            '',
                            '',
                            f"Ошибка: {result.get('error', '')}"
                        ])
            
            return filename
        except Exception as e:
            raise Exception(f"Ошибка при экспорте в CSV: {e}")
    
    def export_stats_to_excel(self, results, filename=None):
        """Экспорт статистики в Excel"""
        if not HAS_OPENPYXL:
            raise Exception("Для экспорта в Excel требуется библиотека openpyxl. Установите: pip install openpyxl")
        
        if filename is None:
            filename = f"stats_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Статистика каналов"
            
            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Заголовки
            headers = ['Канал', 'URL', 'Подписчики', 'Просмотры', 'Видео', 'Дата регистрации', 'Страна', 'Статус']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            for row_idx, result in enumerate(results, 2):
                if 'error' not in result:
                    ws.cell(row=row_idx, column=1, value=result.get('channel_name', ''))
                    ws.cell(row=row_idx, column=2, value=result.get('url', ''))
                    ws.cell(row=row_idx, column=3, value=result.get('subscribers', '0'))
                    ws.cell(row=row_idx, column=4, value=result.get('total_views', '0'))
                    ws.cell(row=row_idx, column=5, value=result.get('videos_count', '0'))
                    ws.cell(row=row_idx, column=6, value=result.get('join_date', ''))
                    ws.cell(row=row_idx, column=7, value=result.get('country', ''))
                    ws.cell(row=row_idx, column=8, value='Успешно')
                else:
                    ws.cell(row=row_idx, column=2, value=result.get('url', ''))
                    ws.cell(row=row_idx, column=8, value=f"Ошибка: {result.get('error', '')}")
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Ошибка при экспорте в Excel: {e}")
    
    def export_stats_to_json(self, results, filename=None):
        """Экспорт статистики в JSON"""
        if filename is None:
            filename = f"stats_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        try:
            export_data = {
                'export_date': datetime.now().isoformat(),
                'total_channels': len(results),
                'channels': results
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            return filename
        except Exception as e:
            raise Exception(f"Ошибка при экспорте в JSON: {e}")
    
    def export_accounts_to_csv(self, accounts_data, filename=None):
        """Экспорт аккаунтов в CSV"""
        if filename is None:
            filename = f"accounts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                
                # Заголовки
                writer.writerow(['Аккаунт', 'Папка', 'Материалов', 'Размер', 'Дата изменения', 'Качество'])
                
                # Данные
                for account in accounts_data:
                    writer.writerow([
                        account.get('name', ''),
                        account.get('folder', ''),
                        account.get('materials_count', 0),
                        account.get('size', ''),
                        account.get('modified_date', ''),
                        account.get('quality_score', '')
                    ])
            
            return filename
        except Exception as e:
            raise Exception(f"Ошибка при экспорте аккаунтов: {e}")

